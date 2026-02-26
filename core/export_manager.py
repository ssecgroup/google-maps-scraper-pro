import csv
import json
import pandas as pd
from datetime import datetime
import os
from typing import List, Dict, Any
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from jinja2 import Template
import weasyprint
from colorama import Fore, Style

class ExportManager:
    """Advanced export manager with multiple formats and styling"""
    
    def __init__(self, config: Dict):
        self.config = config
        self.logger = logging.getLogger(__name__)
        self.output_dir = 'output'
        self.create_output_directories()
        
    def create_output_directories(self):
        """Create necessary output directories"""
        directories = ['csv', 'excel', 'json', 'database', 'reports', 'html']
        for dir_name in directories:
            path = os.path.join(self.output_dir, dir_name)
            if not os.path.exists(path):
                os.makedirs(path)
                self.logger.info(f"Created directory: {path}")
    
    def generate_filename(self, base_name: str, extension: str, subdir: str = '') -> str:
        """Generate timestamped filename"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{base_name}_{timestamp}.{extension}"
        
        if subdir:
            return os.path.join(self.output_dir, subdir, filename)
        return os.path.join(self.output_dir, filename)
    
    def export_csv(self, data: List[Dict], base_name: str = 'businesses') -> str:
        """Export to CSV with proper formatting"""
        if not data:
            self.logger.warning("No data to export")
            return None
        
        try:
            filename = self.generate_filename(base_name, 'csv', 'csv')
            
            # Flatten nested structures
            flat_data = self.flatten_data(data)
            
            with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                if flat_data:
                    fieldnames = flat_data[0].keys()
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(flat_data)
            
            self.logger.info(f"{Fore.GREEN}✓ CSV exported: {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"CSV export failed: {str(e)}")
            return None
    
    def export_excel(self, data: List[Dict], base_name: str = 'businesses') -> str:
        """Export to Excel with formatting"""
        if not data:
            self.logger.warning("No data to export")
            return None
        
        try:
            filename = self.generate_filename(base_name, 'xlsx', 'excel')
            
            # Flatten data
            flat_data = self.flatten_data(data)
            df = pd.DataFrame(flat_data)
            
            # Create Excel writer
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Write main data sheet
                df.to_excel(writer, sheet_name='Businesses', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Businesses']
                
                # Style header
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column].width = adjusted_width
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1):
                    for cell in row:
                        cell.border = thin_border
                
                # Add summary sheet
                summary_data = self.create_summary(data)
                summary_df = pd.DataFrame([summary_data])
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Style summary sheet
                summary_sheet = writer.sheets['Summary']
                for col in range(1, len(summary_df.columns) + 1):
                    cell = summary_sheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
            
            self.logger.info(f"{Fore.GREEN}✓ Excel exported: {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"Excel export failed: {str(e)}")
            return None
    
    def export_json(self, data: List[Dict], base_name: str = 'businesses') -> str:
        """Export to JSON with pretty printing"""
        try:
            filename = self.generate_filename(base_name, 'json', 'json')
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            
            self.logger.info(f"{Fore.GREEN}✓ JSON exported: {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"JSON export failed: {str(e)}")
            return None
    
    def export_sqlite(self, data: List[Dict], base_name: str = 'businesses') -> str:
        """Export to SQLite database"""
        try:
            filename = os.path.join(self.output_dir, 'database', f'{base_name}.db')
            
            # Flatten data
            flat_data = self.flatten_data(data)
            df = pd.DataFrame(flat_data)
            
            # Connect to SQLite
            conn = sqlite3.connect(filename)
            
            # Write to database
            df.to_sql('businesses', conn, if_exists='replace', index=False)
            
            # Create indexes for common queries
            cursor = conn.cursor()
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_name ON businesses(name)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_category ON businesses(category)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_rating ON businesses(rating)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_zip ON businesses(zip_code)')
            
            # Create views
            cursor.execute('''
                CREATE VIEW IF NOT EXISTS high_rated AS
                SELECT * FROM businesses WHERE rating >= 4.5
            ''')
            
            cursor.execute('''
                CREATE VIEW IF NOT EXISTS with_email AS
                SELECT * FROM businesses WHERE emails IS NOT NULL AND emails != '[]'
            ''')
            
            conn.commit()
            conn.close()
            
            self.logger.info(f"{Fore.GREEN}✓ SQLite exported: {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"SQLite export failed: {str(e)}")
            return None
    
    def export_html_report(self, data: List[Dict], base_name: str = 'report') -> str:
        """Export interactive HTML report"""
        try:
            filename = self.generate_filename(base_name, 'html', 'reports')
            
            # Create summary statistics
            summary = self.create_summary(data)
            
            # HTML template
            html_template = """
            <!DOCTYPE html>
            <html>
            <head>
                <title>Business Scraping Report</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 20px; }
                    h1 { color: #333; }
                    .summary { background: #f5f5f5; padding: 15px; border-radius: 5px; margin-bottom: 20px; }
                    .stat { display: inline-block; margin-right: 30px; }
                    .stat-value { font-size: 24px; font-weight: bold; color: #4F81BD; }
                    .stat-label { color: #666; }
                    table { border-collapse: collapse; width: 100%; }
                    th { background: #4F81BD; color: white; padding: 10px; text-align: left; }
                    td { padding: 8px; border-bottom: 1px solid #ddd; }
                    tr:hover { background: #f5f5f5; }
                    .rating { color: #FFD700; }
                    .email { color: #4CAF50; }
                    .phone { color: #2196F3; }
                    .website { color: #FF9800; }
                </style>
            </head>
            <body>
                <h1>Google Maps Business Scraper Report</h1>
                
                <div class="summary">
                    <h2>Summary Statistics</h2>
                    <div class="stat">
                        <div class="stat-value">{{ summary.total_businesses }}</div>
                        <div class="stat-label">Total Businesses</div>
                    </div>
                    <div class="stat">
                        <div class="stat-value">{{ summary.with_phone }}%</div>
                        <div class="stat-label">With Phone</div>
                    </div>
                    <div class="stat">
                        <div class="stat-value">{{ summary.with_email }}%</div>
                        <div class="stat-label">With Email</div>
                    </div>
                    <div class="stat">
                        <div class="stat-value">{{ summary.with_website }}%</div>
                        <div class="stat-label">With Website</div>
                    </div>
                    <div class="stat">
                        <div class="stat-value">{{ summary.avg_rating }}</div>
                        <div class="stat-label">Avg Rating</div>
                    </div>
                </div>
                
                <h2>Business Details</h2>
                <table>
                    <thead>
                        <tr>
                            <th>#</th>
                            <th>Name</th>
                            <th>Category</th>
                            <th>Rating</th>
                            <th>Phone</th>
                            <th>Email</th>
                            <th>Website</th>
                            <th>Address</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for business in data %}
                        <tr>
                            <td>{{ loop.index }}</td>
                            <td>{{ business.name }}</td>
                            <td>{{ business.category }}</td>
                            <td class="rating">{{ business.rating }} ({{ business.reviews }})</td>
                            <td class="phone">{{ business.primary_phone or 'N/A' }}</td>
                            <td class="email">{{ business.primary_email or 'N/A' }}</td>
                            <td class="website"><a href="{{ business.website }}">Link</a></td>
                            <td>{{ business.address }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
                
                <script>
                    // Add sorting functionality
                    document.querySelectorAll('th').forEach(header => {
                        header.addEventListener('click', () => {
                            const table = header.closest('table');
                            const index = Array.from(header.parentNode.children).indexOf(header);
                            const rows = Array.from(table.querySelectorAll('tbody tr'));
                            
                            rows.sort((a, b) => {
                                const aVal = a.children[index].textContent;
                                const bVal = b.children[index].textContent;
                                return aVal.localeCompare(bVal);
                            });
                            
                            rows.forEach(row => table.querySelector('tbody').appendChild(row));
                        });
                    });
                </script>
            </body>
            </html>
            """
            
            # Render template
            template = Template(html_template)
            html_content = template.render(data=data, summary=summary)
            
            # Save HTML
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.logger.info(f"{Fore.GREEN}✓ HTML report exported: {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"HTML export failed: {str(e)}")
            return None
    
    def export_all(self, data: List[Dict], base_name: str = 'businesses') -> Dict:
        """Export in all configured formats"""
        results = {}
        
        formats = self.config['export_settings']['formats']
        
        if 'csv' in formats:
            results['csv'] = self.export_csv(data, base_name)
        
        if 'excel' in formats:
            results['excel'] = self.export_excel(data, base_name)
        
        if 'json' in formats:
            results['json'] = self.export_json(data, base_name)
        
        if 'sqlite' in formats:
            results['sqlite'] = self.export_sqlite(data, base_name)
        
        if 'html_report' in formats:
            results['html'] = self.export_html_report(data, base_name)
        
        return results
    
    def flatten_data(self, data: List[Dict]) -> List[Dict]:
        """Flatten nested data structures for CSV export"""
        flat_list = []
        
        for item in data:
            flat_item = {}
            
            for key, value in item.items():
                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        flat_item[f"{key}_{sub_key}"] = str(sub_value) if sub_value else ''
                elif isinstance(value, list):
                    if key == 'emails' and value:
                        flat_item[key] = ', '.join(str(v) for v in value)
                    elif key == 'phone_numbers' and value:
                        flat_item[key] = ', '.join(str(v) for v in value)
                    elif value:
                        flat_item[key] = str(value)
                    else:
                        flat_item[key] = ''
                else:
                    flat_item[key] = str(value) if value else ''
            
            flat_list.append(flat_item)
        
        return flat_list
    
    def create_summary(self, data: List[Dict]) -> Dict:
        """Create summary statistics"""
        total = len(data)
        
        if total == 0:
            return {
                'total_businesses': 0,
                'with_phone': 0,
                'with_email': 0,
                'with_website': 0,
                'avg_rating': 0,
                'scrape_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
        
        with_phone = sum(1 for b in data if b.get('phone_numbers'))
        with_email = sum(1 for b in data if b.get('emails'))
        with_website = sum(1 for b in data if b.get('website'))
        
        ratings = [b.get('rating', 0) for b in data if b.get('rating')]
        avg_rating = sum(ratings) / len(ratings) if ratings else 0
        
        return {
            'total_businesses': total,
            'with_phone': round((with_phone / total) * 100, 2),
            'with_email': round((with_email / total) * 100, 2),
            'with_website': round((with_website / total) * 100, 2),
            'avg_rating': round(avg_rating, 2),
            'scrape_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
